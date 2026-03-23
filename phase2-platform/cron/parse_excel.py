#!/usr/bin/env python3
"""
parse_excel.py  — AMZ Retail Platform
Reads 'Amazon FBM' sheet from the uploaded Excel file.
Column mapping is 100% exact to the original spreadsheet headers.
Formulas are evaluated via data_only=True fallback; raw formula text
is also captured for the formula editor in Settings → Formulas.

Usage:
    python3 parse_excel.py <input.xlsx> <output.json>
"""

import sys
import json
import re
import os

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Column definitions — 100% exact header names ─────────────────────────────
# key = JSON field name (snake_case),  value = Excel header (exact)
COLUMN_MAP = {
    "ean_amazon":              "EAN Amazon",
    "ean_dostavchik":          "EAN Доставчик",
    "korekcia_cena":           "Корекция  на цена",
    "komentar":                "Коментар",
    "nashe_sku":               "Наше SKU",
    "dostavchik_sku":          "Доставчик SKU",
    "dostavchik":              "Доставчик",
    "brand":                   "Бранд",
    "model":                   "Модел",
    "amazon_link":             "Amazon Link",
    "asin":                    "ASIN",
    "cena_konkurent":          "Цена Конкурент  - Brutto",
    "cena_amazon":             "Цена Amazon  - Brutto",
    "prodazhna_cena":          "Продажна Цена в Амазон  - Brutto",
    "cena_bez_dds":            "Цена без ДДС",
    "dds_prodazhna":           "ДДС от продажна цена",
    "amazon_taksi":            "Amazon Такси",
    "cena_dostavchik":         "Цена Доставчик -Netto",
    "dds_dostavchik":          "ДДС  от Цена Доставчик",
    "transport_ot_dostavchik": "Транспорт от Доставчик до нас",
    "transport_do_klient":     "Транспорт до кр. лиент  Netto",
    "dds_transport":           "ДДС  от Транспорт до кр. лиент",
    "rezultat":                "Резултат",
    "namerena_2ra":            "Намерена 2ра обява",
    "cena_ispania":            "Цена за Испания / Франция / Италия",
    "dm_cena":                 "DM цена",
    "nova_cena":               "Нова цена след намаление",
    "dostaveni":               "Доставени",
    "za_sledvashta_poruchka":  "За следваща поръчка",
    "elektronika":             "Електоника",
}

# Columns that have formulas (exact header names)
FORMULA_COLUMNS = {
    "Цена без ДДС":                        "=N/1.19",
    "ДДС от продажна цена":                "=N-O",
    "Amazon Такси":                        "=N*0.15",
    "ДДС  от Цена Доставчик":              "=R*0.2",
    "ДДС  от Транспорт до кр. лиент":     "=U*0.2",
    "Резултат":                            "=N-P-Q-R-T-U+S+V",
    "Цена за Испания / Франция / Италия":  "=N+1.5",
}


def col_letter_to_name(letter, col_by_letter):
    """Map a column letter (N, O, P…) to a header name."""
    return col_by_letter.get(letter.upper(), letter)


def humanize_formula(raw_formula, col_by_letter):
    """Replace Excel column letters with column header names in a formula."""
    def replacer(m):
        letter = m.group(1)
        name = col_letter_to_name(letter, col_by_letter)
        return "{" + name + "}"
    return re.sub(r'\b([A-Z]{1,2})\d+\b', lambda m: "{" + col_letter_to_name(m.group(1), col_by_letter) + "}", raw_formula)


def safe_float(v):
    if v is None:
        return None
    try:
        f = float(v)
        return round(f, 4) if f != int(f) else int(f)
    except (ValueError, TypeError):
        return None


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def load_workbook_values(path):
    """Load workbook with data_only=True to get computed values."""
    return openpyxl.load_workbook(path, data_only=True)


def load_workbook_formulas(path):
    """Load workbook with data_only=False to get raw formulas."""
    return openpyxl.load_workbook(path, data_only=False)


def parse(input_path, output_path):
    if not os.path.exists(input_path):
        print(f"ERROR: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading (values): {input_path}")
    wb_vals = load_workbook_values(input_path)
    print(f"Loading (formulas): {input_path}")
    wb_fmls = load_workbook_formulas(input_path)

    sheet_name = "Amazon FBM"
    if sheet_name not in wb_vals.sheetnames:
        # Try first sheet
        sheet_name = wb_vals.sheetnames[0]
        print(f"WARNING: 'Amazon FBM' not found, using '{sheet_name}'")

    ws_vals = wb_vals[sheet_name]
    ws_fmls = wb_fmls[sheet_name]

    # ── Read headers from row 1 ────────────────────────────────────────────
    header_to_col = {}   # header_name → column_index (1-based)
    col_to_header = {}   # column_index → header_name
    col_letter_map = {}  # column_letter → header_name

    for cell in ws_vals[1]:
        if cell.value is not None:
            h = str(cell.value).strip()
            header_to_col[h] = cell.column
            col_to_header[cell.column] = h
            col_letter_map[cell.column_letter] = h

    print(f"Headers found: {list(header_to_col.keys())}")

    # Build reverse map: JSON field → column index
    field_to_col = {}
    for field, header in COLUMN_MAP.items():
        if header in header_to_col:
            field_to_col[field] = header_to_col[header]
        else:
            # Try case-insensitive / whitespace-normalised match
            for h, ci in header_to_col.items():
                if h.replace("  ", " ").strip().lower() == header.replace("  ", " ").strip().lower():
                    field_to_col[field] = ci
                    print(f"  Fuzzy match: '{field}' → col {ci} ('{h}')")
                    break

    print(f"Mapped fields: {len(field_to_col)}/{len(COLUMN_MAP)}")

    # ── Extract formula templates from row 2 (using formula sheet) ─────────
    formula_templates = {}
    for cell in ws_fmls[2]:
        v = cell.value
        if v and str(v).startswith("="):
            hdr = col_to_header.get(cell.column, "")
            if hdr:
                # Convert e.g. =N2/1.19 → =N/1.19 (strip row number)
                clean = re.sub(r'([A-Z]+)\d+', r'\1', str(v))
                formula_templates[hdr] = clean

    print(f"Formula templates extracted: {formula_templates}")

    # ── Parse data rows ────────────────────────────────────────────────────
    products = []
    rows_vals = list(ws_vals.iter_rows(min_row=2, values_only=True))

    for row_vals in rows_vals:
        # Skip completely empty rows
        if all(v is None for v in row_vals):
            continue

        # Helper to get value by column index (1-based) from this row tuple
        def get(col_idx):
            if col_idx is None:
                return None
            idx = col_idx - 1
            if idx < 0 or idx >= len(row_vals):
                return None
            return row_vals[idx]

        def get_str(field):
            ci = field_to_col.get(field)
            return safe_str(get(ci))

        def get_num(field):
            ci = field_to_col.get(field)
            return safe_float(get(ci))

        def get_bool(field, true_val="Yes"):
            ci = field_to_col.get(field)
            v = get(ci)
            if v is None:
                return ""
            s = str(v).strip().lower()
            if s in ("yes", "true", "1", "да", "y"):
                return "Yes"
            if s in ("no", "false", "0", "не", "n"):
                return "No"
            return safe_str(v)

        # Must have at least EAN or ASIN
        ean = get_str("ean_amazon")
        asin = get_str("asin")
        if not ean and not asin:
            continue

        product = {
            # ── Identifiers ──────────────────────────────────────────
            "EAN Amazon":              ean,
            "EAN Доставчик":           get_str("ean_dostavchik"),
            "Корекция  на цена":       get_str("korekcia_cena"),
            "Коментар":                get_str("komentar"),
            "Наше SKU":                get_str("nashe_sku"),
            "Доставчик SKU":           get_str("dostavchik_sku"),
            "Доставчик":               get_str("dostavchik"),
            "Бранд":                   get_str("brand"),
            "Модел":                   get_str("model"),
            "Amazon Link":             get_str("amazon_link"),
            "ASIN":                    asin,
            # ── Prices ───────────────────────────────────────────────
            "Цена Конкурент  - Brutto":  get_num("cena_konkurent"),
            "Цена Amazon  - Brutto":     get_num("cena_amazon"),
            "Продажна Цена в Амазон  - Brutto": get_num("prodazhna_cena"),
            "Цена без ДДС":              get_num("cena_bez_dds"),
            "ДДС от продажна цена":      get_num("dds_prodazhna"),
            "Amazon Такси":              get_num("amazon_taksi"),
            "Цена Доставчик -Netto":     get_num("cena_dostavchik"),
            "ДДС  от Цена Доставчик":    get_num("dds_dostavchik"),
            "Транспорт от Доставчик до нас": get_num("transport_ot_dostavchik"),
            "Транспорт до кр. лиент  Netto": get_num("transport_do_klient"),
            "ДДС  от Транспорт до кр. лиент": get_num("dds_transport"),
            "Резултат":                  get_num("rezultat"),
            # ── Other ────────────────────────────────────────────────
            "Намерена 2ра обява":        get_str("namerena_2ra"),
            "Цена за Испания / Франция / Италия": get_num("cena_ispania"),
            "DM цена":                   get_num("dm_cena"),
            "Нова цена след намаление":  get_num("nova_cena"),
            "Доставени":                 get_num("dostaveni"),
            "За следваща поръчка":       get_num("za_sledvashta_poruchka"),
            "Електоника":                get_bool("elektronika"),
            # ── Internal fields ──────────────────────────────────────
            "_upload_status":            "NOT_UPLOADED",
            "_source":                   "Amazon FBM",
        }
        products.append(product)

    print(f"Parsed {len(products)} products")

    # ── Build output ──────────────────────────────────────────────────────
    output = {
        "products": products,
        "columns": list(COLUMN_MAP.values()),      # ordered column list (exact names)
        "formula_templates": formula_templates,     # column header → formula template
        "meta": {
            "total": len(products),
            "sheet": sheet_name,
            "source_file": os.path.basename(input_path),
        }
    }

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Saved to: {output_path} ({os.path.getsize(output_path)//1024} KB)")
    return len(products)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <input.xlsx> <output.json>")
        sys.exit(1)
    n = parse(sys.argv[1], sys.argv[2])
    print(f"Done: {n} products exported")
